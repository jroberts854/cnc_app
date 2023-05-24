import requests
from openpyxl import Workbook
from openpyxl.styles import Font
import pprint
from orders import calc
import datetime

# import pandas as pd


url = "https://uloll00m1i.execute-api.ca-central-1.amazonaws.com/master/"
response = requests.get(url)
data = response.json()


# Receive order id's and return order info
def get_products(order_id):
    # fetch from api
    response = requests.get(f"{url}/orders/{order_id}")
    order_data = response.json()
    ready_date = order_data["ready_date"]
    orderproducts = order_data["quote"]["orderproducts"]
    job_title = order_data["quote"]["project"]["job_title"]
    client = order_data["quote"]["project"]["company_name"]
    title = f'#{order_id}-{job_title}-{client}'

    # Fetch product infomation here
    order_data = {}
    order_data_list = []
    for products in orderproducts:
        model = products["model"]
        qty = products["quantity"]
        products_options = products["options"]
        options = []
        for option in products_options:
            option_code = option['code']
            option_type = option["option_type"]
            options.append([option_type, option_code])
        order_data_list.append([model, qty, options])
        order_data[order_id] = [title,ready_date, order_data_list]

    # pprint.pprint(order_data)
    return order_data


# Receive product data and returns 1 calc per interation
def get_cnc_calc(order_data):

    cutlist_data = {}
    calc_data = []
    for order_id, products in order_data.items():
        job_title = products[0]
        ready_date = products[1]
        for product in products[2]:
            model = product[0]
            qty = product[1]
            options = product[2]
            cnc_data = calc(model, qty, options)
            calc_data.append([model, cnc_data.split("\n")])

            # cutlist_data[order_id] = calc_data
            # pprint.pprint(cutlist_data)

    return [job_title,ready_date,calc_data]



# orders to xslx file
# orders data format
# {
#     order_num1: [[model1,calcs],[model2, calcs]],
#     order_num2: [[model3,calcs],[model4, calcs]]
# }

def orders_to_xslx(orders, filename):

    # make workbook
    workbook = Workbook()
    # # make sheets
    # # ws_joblist = workbook.create_sheet("JOB-LIST")
    # Workbook.active
    # ws_cutlist = workbook.create_sheet("CUTLIST")

    ws_joblist = workbook.active
    ws_joblist.title = 'JOB-LIST'
    ws_cutlist = workbook.create_sheet('CUTLIST')



    # get active worksheet
    # ws_joblist = workbook.active
    # ws_cutlist = workbook.active
################# JOBLIST ########################################
    #JOB LIST
    ws_joblist["A1"] = "JOB"
    ws_joblist["B1"] = "READY DATE"
    ws_joblist["C1"] = "NOTES"

    row_index=2
    for order_num ,models in orders.items():

        title = models[0]
        ready_date = models[1]
        ready_date = ready_date[0:10]

        cell = ws_joblist["A" + str(row_index)]
        cell.value = "JOB" + title

        cell = ws_joblist["B" + str(row_index)]
        cell.value = ready_date
        cell.font = Font(bold=False)
        row_index += 1





################# CUSTLIST #######################################

    # write out columns
    ws_cutlist["A1"] = "ORDER_NUM"
    ws_cutlist["B1"] = "WORKPEICE NAME"
    ws_cutlist["C1"] = "TYPE"
    ws_cutlist["D1"] = "DESCRIPTION "
    ws_cutlist["E1"] = "QTY"
    ws_cutlist["F1"] = "XQTY"
    ws_cutlist["G1"] = "X"
    ws_cutlist["H1"] = "Y"
    ws_cutlist["I1"] = "Z"
    ws_cutlist["J1"] = "MATERIAL"
    ws_cutlist["K1"] = "GRAIN"
    ws_cutlist["L1"] = "FILE"
    ws_cutlist["M1"] = "STATUS"

    # start at row 3
    row_index = 3

    # for each order
    for order_num ,models in orders.items():

        # display order vumber
        print(order_num)
        cell = ws_cutlist["B" +str(row_index)]
        # cell.value = 'order number# '+ str(order_num)
        title = models[0]
        cell.value = title
        cell.font = Font(bold=True)
        row_index += 2

        # for each model and calcs
        for model_calcs in models[2]:
            # get model number
            model = model_calcs[0]
            # get calcs for this model
            calcs = model_calcs[1]

            # display order number bold
            print(order_num)
            cell = ws_cutlist["A" +str(row_index)]
            cell.value = int(order_num)
            cell.font = Font(bold=True)

            # display model bold
            print(model)
            cell = ws_cutlist["B" +str(row_index)]
            cell.value = "model" + model
            cell.font = Font(bold=True)
            row_index += 1

            # for calcs in model calcs
            for calc in calcs:

                # model not foubd
                if 'Model not found' in calc:
                    ws_cutlist.cell(row=row_index, column=2, value='Model not found')
                    row_index +=1

                # write out calcs for model
                else:

                    # convert string to a list
                    row_data = calc.split(";")


                    # check for empty row
                    if len(row_data ) >1:

                        #print(row_data)

                        # insert the new row at the specified row index
                        ws_cutlist.insert_rows(row_index)

                        # insert order number
                        ws_cutlist.cell(row=row_index, column=1, value=int(order_num))

                        # write the new row data to the worksheet
                        for col, value in enumerate(row_data, start=1):

                            # type int
                            if col + 1 == 3:
                                ws_cutlist['C' +str(row_index) ] =int(value)
                            # qty int
                            elif col + 1 == 5:
                                ws_cutlist['E' +str(row_index) ] =int(value)
                            # xqty int
                            elif col + 1 == 6:
                                ws_cutlist['F' +str(row_index) ] =int(value)
                            # x float 3 decimal places
                            elif col + 1 == 7:
                                ws_cutlist['G' +str(row_index)].number_format ='0.000'
                                ws_cutlist['G' +str(row_index) ] =float(value)
                            # y float 3 decimal places
                            elif col + 1 == 8:
                                ws_cutlist['H' +str(row_index)].number_format ='0.000'
                                ws_cutlist['H' +str(row_index) ] =float(value)
                            # z float 2 decimal places
                            elif col + 1 == 9:
                                ws_cutlist['I' +str(row_index)].number_format ='0.00'
                                ws_cutlist['I' +str(row_index) ] =float(value)
                            # frain int
                            elif col + 1 == 11:
                                ws_cutlist['K' +str(row_index) ] =int(value)

                            # non numeric
                            else:
                                # write non-numeric data
                                ws_cutlist.cell(row=row_index, column=col +1, value=value)

                        row_index +=1 # next line

            # row spacre between calcs
            row_index +=1

    # save workbook
    print("wrote to workbook ", filename)
    workbook.save(filename=filename)



if __name__ == '__main__':
    # order_data(7530)
    order_list = [7530, 7529]
    for order in order_list:
        # get_products(order)
        result = get_cnc_calc(get_products(order))
    # pprint.pprint(result)
    cutlist = orders_to_xslx(result, "cutlist1.xlsx")
